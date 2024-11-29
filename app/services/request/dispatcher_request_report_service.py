import io
import textwrap
import zipfile
from collections.abc import AsyncGenerator, AsyncIterable
from datetime import datetime

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

from client.s300.models.employee import EmployeeS300
from config import settings
from constants import MONTHS
from models.request.categories_tree import (
    REQUEST_CATEGORY_EN_RU,
    REQUEST_SUBCATEGORY_EN_RU,
)
from models.request.constants import REQUEST_SOURCE_EN_RU
from models.request.embs.requester import RequesterType
from models.request.request import RequestModel

pdfmetrics.registerFont(TTFont("Roboto", f"{settings.PROJECT_DIR}/static/fonts/Roboto/Roboto-Regular.ttf"))
pdfmetrics.registerFont(TTFont("RobotoBold", f"{settings.PROJECT_DIR}/static/fonts/Roboto/Roboto-Bold.ttf"))


class DispatcherRequestReportService:
    """
    Сервис для работы с отчетами по заявкам для сотрудников
    """

    def __init__(self, employee: EmployeeS300):
        """
        Инициализация сервиса

        Args:
            employee (EmployeeS300): Модель работника осуществляющего работу с позициями каталога
        """

        super().__init__()
        self.employee = employee

    async def generate_xlsx_report(self, requests: AsyncIterable[RequestModel]) -> AsyncGenerator[bytes, None]:
        wb = Workbook()
        ws = wb.active
        if not ws:
            raise Exception("Failed to create a Workbook")
        ws.title = f"Журнал заявок, {self.employee.provider.name}.xlsx"
        ws.merge_cells("A1:C1")
        ws["A1"] = f"Журнал заявок,\n{self.employee.provider.name}\nОтчет сформирован {datetime.now().strftime("%d.%m.%Y %H:%M")}"
        ws.append(
            [
                "Дата",
                "Номер",
                "Источник",
                "Адрес",
                "ФИО заявителя",
                "Телефон заявителя",
                "Категория",
                "Подкатегория",
                "Описание",
                "Комментарий к выполненным работам",
                "Желаемое время выполнения",
                "Дата исполнения",
                "Исполнители",
                "Смета",
                "Итоговая стоимость",
            ]
        )
        async for request in requests:
            address = request.house.address
            if request.area:
                address = address + request.area.formatted_number
            total_cost = 0
            estimate = ""
            if request.resources.warehouses:
                estimate += "Материалы со складов:"
                for wh in request.resources.warehouses:
                    for i in wh.items:
                        cost = i.price * i.quantity
                        total_cost += cost
                        estimate += f"\n  {i.name} Количество: {i.quantity}. Цена: {i.price}. Стоимость: {cost}"
            if request.commerce.catalog_items:
                if estimate:
                    estimate += "\n"
                estimate += "Позиции каталога:"
                for ci in request.commerce.catalog_items:
                    cost = ci.price * ci.quantity
                    total_cost += cost
                    estimate += f"\n  {ci.name} Количество: {ci.quantity}. Цена: {ci.price}. Стоимость: {cost}"
            if request.resources.services:
                if estimate:
                    estimate += "\n"
                estimate += "Вручную добавленные услуги:"
                for s in request.resources.services:
                    cost = s.price * s.quantity
                    total_cost += cost
                    estimate += f"\n  {s.name} Количество: {s.quantity}. Цена: {s.price}. Стоимость: {cost}"
            if request.resources.services:
                if estimate:
                    estimate += "\n"
                estimate += "Вручную добавленные материалы:"
                for m in request.resources.materials:
                    cost = m.price * m.quantity
                    total_cost += cost
                    estimate += f"\n  {m.name} Количество: {m.quantity}. Цена: {m.price}. Стоимость: {cost}"
            ws.append(
                [
                    request.created_at.strftime("%d.%m.%Y"),
                    request.number,
                    REQUEST_SOURCE_EN_RU[request.source],
                    address,
                    request.requester.full_name,
                    "\n".join([f"+7 {ph.number} {ph.add}" for ph in request.requester.phone_numbers]),
                    REQUEST_CATEGORY_EN_RU[request.category],
                    REQUEST_SUBCATEGORY_EN_RU[request.subcategory] if request.subcategory else None,
                    request.description,
                    request.execution.description,
                    request.execution.desired_end_at,
                    request.execution.end_at,
                    "\n".join(e.full_name for e in request.execution.employees),
                    estimate,
                    total_cost,
                ]
            )

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        chunk_size = 1024
        while chunk := buffer.read(chunk_size):
            yield chunk
        buffer.close()

    async def generate_pdf_blanks_zip(self, requests: AsyncIterable[RequestModel]) -> AsyncGenerator[bytes, None]:
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            async for request in requests:
                pdf_buffer = io.BytesIO()
                async for chunk in self.generate_pdf_blank(request):
                    pdf_buffer.write(chunk)
                pdf_buffer.seek(0)
                zf.writestr(f"Бланк заявки №{request.number}.pdf", pdf_buffer.read())

        buffer.seek(0)
        chunk_size = 1024
        while chunk := buffer.read(chunk_size):
            yield chunk
        buffer.close()

    async def generate_pdf_blank(self, request: RequestModel) -> AsyncGenerator[bytes, None]:

        # Создание PDF-документа
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        # Заголовок
        current_height = 10
        c.setFont("RobotoBold", 12)
        c.drawString(10 * mm, height - current_height * mm, request.provider.name)

        # Номер заявки и дата
        c.setFont("RobotoBold", 18)
        current_height += 20
        c.drawString(65 * mm, height - current_height * mm, f"ЗАЯВКА №{request.number}")
        c.setFont("RobotoBold", 16)
        current_height += 10
        c.drawString(70 * mm, height - current_height * mm, f"от «{request.created_at.strftime("%d")}» {MONTHS[request.created_at.month]["ru_genitive"]} {request.created_at.year}г.")
        c.setFont("RobotoBold", 12)

        # Основной блок с данными о заказчике
        current_height += 15
        c.drawString(10 * mm, height - current_height * mm, "Информация о заявителе:")
        c.setFont("Roboto", 10)
        current_height += 7
        c.setFont("RobotoBold", 10)
        c.drawString(15 * mm, height - current_height * mm, "ФИО:")
        c.setFont("Roboto", 10)
        c.drawString(25 * mm, height - current_height * mm, request.requester.full_name[:87])
        current_height += 5
        c.setFont("RobotoBold", 10)
        c.drawString(15 * mm, height - current_height * mm, "Телефон:")
        c.setFont("Roboto", 10)
        phone_numbers = ", ".join([f'+7 ({ph.number[:3]}) {ph.number[3:]}{" доб. " + ph.add if ph.add else ""}' for ph in request.requester.phone_numbers[:5]])
        c.drawString(35 * mm, height - current_height * mm, phone_numbers[:99])
        current_height += 5
        if request.requester.type == RequesterType.TENANT:
            c.setFont("RobotoBold", 10)
            c.drawString(15 * mm, height - current_height * mm, "Адрес проживания:")
            address = request.requester.house.address + ", " + request.requester.area.formatted_number
            c.setFont("Roboto", 9 if len(address) < 96 else 8 if len(address) < 106 else 7)
            c.drawString(50 * mm, height - current_height * mm, address[:120])
        current_height += 5
        c.rect(5 * mm, height - current_height * mm, width - 10 * mm, 30 * mm, stroke=1, fill=0)

        c.setFont("RobotoBold", 12)
        current_height += 10
        c.drawString(10 * mm, height - current_height * mm, "Адрес:")
        current_height += 7
        address = request.house.address + (", " + request.area.formatted_number if request.area else "")
        c.setFont("Roboto", 10 if len(address) < 106 else 9 if len(address) < 117 else 8)
        c.drawString(15 * mm, height - current_height * mm, address[:129])
        c.setFont("RobotoBold", 12)
        current_height += 10
        c.drawString(10 * mm, height - current_height * mm, "Описание:")
        c.setFont("Roboto", 10)
        current_height += 2
        description_chunks = textwrap.wrap(request.description, width=100)
        if len(description_chunks) > 5:
            description_chunks = description_chunks[:5]
            description_chunks.append("(Полное описание заявки доступно на сайте)")
        for line in description_chunks:
            current_height += 5
            c.drawString(15 * mm, height - current_height * mm, line)
        current_height += 10
        c.setFont("RobotoBold", 12)
        c.drawString(10 * mm, height - current_height * mm, "Желаемое время выполнения:")
        c.drawString(120 * mm, height - current_height * mm, "Примечания:")
        c.setFont("Roboto", 10)
        current_height += 7
        c.drawString(15 * mm, height - current_height * mm, f"с    {request.execution.desired_start_at.strftime("%H:%M  %d.%m.%Y") if request.execution.desired_start_at else "не указано"}")
        c.line(125 * mm, height - current_height * mm, 188 * mm, height - current_height * mm)
        current_height += 5
        c.drawString(15 * mm, height - current_height * mm, f"по  {request.execution.desired_end_at.strftime("%H:%M  %d.%m.%Y") if request.execution.desired_end_at else "не указано"}")
        c.line(125 * mm, height - current_height * mm, 188 * mm, height - current_height * mm)

        current_height += 10
        c.setFont("RobotoBold", 12)
        c.drawString(10 * mm, height - current_height * mm, "Фактическое время выполнения:")
        c.drawString(120 * mm, height - current_height * mm, "Исполнители:")
        c.setFont("Roboto", 10)
        current_height += 7
        c.drawString(15 * mm, height - current_height * mm, "       :            «        »                       20       г.")
        c.line(15 * mm, height - current_height * mm, 20 * mm, height - current_height * mm)
        c.line(23 * mm, height - current_height * mm, 28 * mm, height - current_height * mm)
        c.line(35 * mm, height - current_height * mm, 40 * mm, height - current_height * mm)
        c.line(45 * mm, height - current_height * mm, 60 * mm, height - current_height * mm)
        c.line(67 * mm, height - current_height * mm, 72 * mm, height - current_height * mm)
        c.line(125 * mm, height - current_height * mm, 188 * mm, height - current_height * mm)
        current_height += 5
        c.line(125 * mm, height - current_height * mm, 188 * mm, height - current_height * mm)
        current_height += 5
        c.setFont("RobotoBold", 12)
        c.drawString(10 * mm, height - current_height * mm, "Выполненные работы:")
        c.setFont("Roboto", 10)
        for _ in range(5):
            current_height += 5
            c.line(15 * mm, height - current_height * mm, 188 * mm, height - current_height * mm)

        current_height += 10
        c.setFont("RobotoBold", 12)
        c.drawString(10 * mm, height - current_height * mm, "Затраченные материалы:")
        c.setFont("Roboto", 10)
        for _ in range(3):
            current_height += 5
            c.line(15 * mm, height - current_height * mm, 188 * mm, height - current_height * mm)
            c.setFont("RobotoBold", 12)
        current_height += 10
        c.drawString(10 * mm, height - current_height * mm, "Квартирные счетчики(номер, тариф, показания):")
        c.setFont("Roboto", 10)
        for _ in range(3):
            current_height += 5
            c.line(15 * mm, height - current_height * mm, 188 * mm, height - current_height * mm)

        current_height += 10
        c.setFont("RobotoBold", 12)
        c.drawString(10 * mm, height - current_height * mm, "Работа выполнена:")
        c.setFont("Roboto", 10)
        current_height += 7
        c.line(15 * mm, height - current_height * mm, 78 * mm, height - current_height * mm)
        current_height += 4
        c.setFont("Roboto", 8)
        c.drawString(35 * mm, height - current_height * mm, "(да, нет, частично)")
        c.save()

        buffer.seek(0)
        chunk_size = 1024
        while chunk := buffer.read(chunk_size):
            yield chunk
        buffer.close()
